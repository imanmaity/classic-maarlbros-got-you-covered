#!/usr/bin/env python3
"""Fetch from the admin's email via IMAP:
  1) the newest schedule .xlsx attachment  -> argv[1] (default rosters/schedule_latest.xlsx)
  2) recent "Change in class schedule" notices -> data/changes.json (parsed)

Env vars (set as GitHub repository secrets):
  MAIL_USER, MAIL_PASS, SENDER, IMAP_HOST (default imap.gmail.com),
  MAIL_SUBJECT (optional, required substring in the schedule email's subject),
  CHANGE_SUBJECT (optional, default "change in class")
Exits non-zero if no schedule attachment is found (so a stale week is never published).
Change parsing is best-effort and never fails the build.
"""
import imaplib, email, os, sys, re, json, datetime
from email.header import decode_header
from email.utils import parsedate_to_datetime, parseaddr

OUT  = sys.argv[1] if len(sys.argv) > 1 else "rosters/schedule_latest.xlsx"
CHANGES_OUT = os.environ.get("CHANGES_OUT", "data/changes.json")
HOST = os.environ.get("IMAP_HOST", "imap.gmail.com")
USER = os.environ.get("MAIL_USER"); PWD = os.environ.get("MAIL_PASS")
SENDER = os.environ.get("SENDER", "mba.im@nirmauni.ac.in")
SUBJ = os.environ.get("MAIL_SUBJECT", "")
CHANGE_SUBJECT = os.environ.get("CHANGE_SUBJECT", "change in class")
if not (USER and PWD):
    sys.exit("MAIL_USER / MAIL_PASS not set.")

def decode(s):
    return "".join(p.decode(enc or "utf-8", "ignore") if isinstance(p, bytes) else p
                   for p, enc in decode_header(s or ""))

def body_text(msg):
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain" and not part.get_filename():
                try: return part.get_payload(decode=True).decode(part.get_content_charset() or "utf-8", "ignore")
                except Exception: pass
        for part in msg.walk():
            if part.get_content_type() == "text/html":
                html = part.get_payload(decode=True).decode("utf-8", "ignore")
                return re.sub(r"<[^>]+>", " ", html)
        return ""
    try: return msg.get_payload(decode=True).decode(msg.get_content_charset() or "utf-8", "ignore")
    except Exception: return ""

# ---- trust guards ----
def _addr_of(msg):
    return parseaddr(decode(msg.get("From", "")))[1].strip().lower()

def _sender_ok(msg):
    if not SENDER:
        return True
    addr, want = _addr_of(msg), SENDER.strip().lower()
    if not addr:
        return False
    if "@" in want:
        return addr == want
    return addr.split("@")[-1] == want            

def _name_dates(fn):
    out = []
    for d, m, y in re.findall(r'(\d{1,2})[._\-/](\d{1,2})[._\-/](\d{2,4})', fn or ""):
        y = int(y); y = 2000 + y if y < 100 else y
        try: out.append(datetime.date(y, int(m), int(d)))
        except ValueError:
            try: out.append(datetime.date(y, int(d), int(m)))
            except ValueError: pass
    return out

def _looks_like_schedule(raw):
    try:
        from openpyxl import load_workbook
        import io
    except Exception:
        return None
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheets = {s: [list(r) for r in wb[s].iter_rows(values_only=True)]
                  for s in wb.sheetnames}
        wb.close()
    except Exception:
        return False
    cd = next((s for s in sheets if "course detail" in s.lower()),
              next(iter(sheets), None))
    cd_rows = sheets.get(cd, [])
    has_codes = any(
        len(r) > 1 and isinstance(r[1], str) and "total" not in r[1].lower()
        and re.search(r"[A-Za-z]", r[1]) and re.search(r"\d", r[1])
        for r in cd_rows[1:])
    has_grid = any(
        any(len(r) > 0 and isinstance(r[0], datetime.datetime) for r in rows[2:])
        for s, rows in sheets.items() if s != cd)
    return bool(has_codes and has_grid)

ROOM_RE = r'(?:[A-Za-z]{1,4}-?\d{1,3}[A-Za-z]?|\d{2,4}-[A-Za-z]{1,2}|\d{3}\s[A-Za-z](?![A-Za-z])|\d{3}[A-Za-z]?)'
_WEEKDAYS = {"monday":0,"tuesday":1,"wednesday":2,"thursday":3,"friday":4,"saturday":5,"sunday":6}

def _room_norm(s):
    s = re.sub(r'\s+', '', str(s)).upper()                 
    s = re.sub(r'^(\d{2,4})([A-Z]{1,2})$', r'\1-\2', s)    
    return s

def _venue_label(s):
    s = str(s).strip().lower()
    if 'zoom' in s:  return 'Zoom'
    if 'webex' in s: return 'Webex'
    if 'meet' in s:  return 'Google Meet'
    if 'team' in s:  return 'MS Teams'
    if re.search(r'online|virtual|classroom', s): return 'Online'
    return re.sub(r'\s+', ' ', s).title()

def _venue_events(t):
    online = (r'(?:online|virtually|virtual\s+mode|ms\s*-?\s*teams|microsoft\s*teams|'
              r'google\s*meet|g-?meet|zoom(?:\s*(?:call|meeting|link))?|webex|google\s*classroom)')
    hall   = (r'(?:auditorium|amphitheat(?:re|er)|seminar\s*hall|conference\s*(?:room|hall)|'
              r'board\s*room|(?:computer\s*)?lab(?:oratory)?|library)')
    cue = (r'(?:held|conducted|conduct\w*|shifted|moved|take\s*place|takes?\s*place|'
           r'venue|be\s+held|arranged)\b[^.\n]{0,25}?\b(?:in|to|at|via|on|through|:)\s*(?:the\s+|a\s+)?')
    ev = []
    for m in re.finditer(cue + r'(' + online + r'|' + hall + r')', t, re.I):
        ev.append((m.start(), _venue_label(m.group(1)), None))
    for m in re.finditer(r'\b(' + online + r')\b', t, re.I):   
        ev.append((m.start(), _venue_label(m.group(1)), None))
    return ev

_CHANGE_SUBJECT_RE = re.compile(
    r'postpon|prepon|reschedul|cancel|class\s*-?\s*room|classroom|\bvenue\b|\broom\b|'
    r'change\s+in\s+(?:class|schedule|time|timing|venue)', re.I)

def _divs(group):
    g = (group or "").strip()
    if re.search(r'\b(?:all|both|entire|every|each)\b', g, re.I):
        return [""]
    return [d.upper() for d in re.findall(r'(?<![A-Za-z.])([A-Ha-h])(?![A-Za-z.])', g)]

def _bare_codes(text):
    stop = {"MBA", "IMBA", "BTECH", "IIM", "PDF", "FYI", "TBA", "AM", "PM",
            "LH", "NOTE", "ALL", "AND", "FOR", "THE", "ARE", "IS"}
    t = text
    while re.search(r'\([^()]*\)', t):
        t = re.sub(r'\([^()]*\)', ' ', t)
    out = []
    def add(a):
        a = a.upper()
        if a not in stop and a not in out:
            out.append(a)
    for ab in re.findall(r'\b([A-Z][A-Z&]{1,5})\b\s+(?i:sessions?|classes?|lectures?|scheduled|additional)\b', t):
        add(ab)                                   
    for sent in re.split(r'\n|(?<!\d)\.(?!\d)', t):   
        if re.search(r'(?i:postpon\w*|prepon\w*|reschedul\w*|cancel\w*|not\s+be\s+held|revis\w*|additional)', sent):
            for ab in re.findall(r'\b([A-Z][A-Z&]{1,5})\b', sent):   
                add(ab)
    return out

def _secs_in(t):
    s = []
    for ab, dvgroup in re.findall(r'([A-Za-z&]{2,6})\(\s*([A-Za-z][A-Za-z&,\s]*?)\s*\)', t):
        for dv in _divs(dvgroup):
            s.append((ab, dv))
    return s

def _detect_rooms(t):
    nr = orr = None
    m = re.search(r'\b(' + ROOM_RE + r')\s+to\s+(' + ROOM_RE + r')\b', t, re.I)   
    if m: orr, nr = _room_norm(m.group(1)), _roomI would be happy to make those changes for you, but I don't have the context of the list or text you are referring to right now. 

Could you please share the list or paste the text here so I can update the 5th item for you?
